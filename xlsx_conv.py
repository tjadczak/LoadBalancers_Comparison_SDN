#!/usr/bin/python3.7
import os, sys, fnmatch
import signal
import re
import fcntl
import array
import struct
import sys
import time
import random
import csv
from openpyxl import Workbook
from openpyxl import load_workbook
import xlcalculator

print("*** XLSX MAGIC ***")
'''for filename in fnmatch.filter(os.listdir('.'), '*.xlsx'):
    results = load_workbook(filename)
    results_sheet = results.active
    for row in range(1,30):
        for column in range(1,12):
            current_value = results_sheet.cell(row=row, column=column).value
            if current_value:
                results_sheet.cell(row=row, column=column, value=float(current_value))
    results_sheet.cell(row=32, column=2, value="=AVERAGE(B1:B30)")
    results_sheet.cell(row=32, column=3, value="=AVERAGE(C1:C30)")
    results_sheet.cell(row=32, column=4, value="=AVERAGE(D1:D30)")
    results_sheet.cell(row=32, column=5, value="=AVERAGE(E1:E30)")
    results_sheet.cell(row=32, column=6, value="=AVERAGE(F1:F30)")
    results_sheet.cell(row=32, column=7, value="=AVERAGE(G1:G30)")
    results_sheet.cell(row=32, column=8, value="=AVERAGE(H1:H30)")
    results_sheet.cell(row=32, column=9, value="=AVERAGE(I1:I30)")
    results_sheet.cell(row=32, column=10, value="=AVERAGE(J1:J30)")
    results_sheet.cell(row=32, column=11, value="=AVERAGE(K1:K30)")

    results_sheet.cell(row=35, column=2, value="=SUM(B32,D32,F32)")
    results_sheet.cell(row=35, column=3, value="=AVERAGE(C32,E32,G32)")
    results_sheet.cell(row=35, column=4, value="=SUM(H32:K32)")

    results.save(filename)'''

main_filename = "Magisterka_dane.xlsx"
main_wb = load_workbook(main_filename)
main_sheet = main_wb.get_sheet_by_name("Results")

files = fnmatch.filter(os.listdir('.'), 'results_params_true_3_random.conf*')
for filename in files:
    num = int(re.findall(r"(\d+).xlsx", filename)[0])
    #results_wb = load_workbook(filename)
    #results_wb = load_workbook(filename, data_only=True, read_only=True)
    #results_sheet = results_wb.active
    results_wb=xw.Book(filename)

    #a = results_sheet.cell(row=35, column=2).internal_value
    a = results_wb.sheets['Sheet'].cell(row=35, column=2).value
    b = results_wb.sheets['Sheet'].cell(row=35, column=3).value
    c = results_wb.sheets['Sheet'].cell(row=35, column=4).value
    #b = results_sheet.cell(row=35, column=3).value
    #c = results_sheet.cell(row=35, column=4).value
    print(a,b,c)
    main_sheet.cell(row=2+num, column=40, value=a)
    main_sheet.cell(row=2+num, column=41, value=b)
    main_sheet.cell(row=2+num, column=42, value=c)
    results_wb.close()
    
main_wb.save(main_filename)
    



print("*** DONE ***")


